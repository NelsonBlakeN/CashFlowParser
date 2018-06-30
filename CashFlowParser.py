# System imports
try:
    import sys
    import os
    import logging
    import smtplib
    from datetime import date, datetime
    from collections import defaultdict
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
except Exception as e:
    # Manually log error
    f = open("/home/blake/Documents/logs/expense_report.log", "w+")
    f.write("ERROR importing system libraries:")
    f.write("{}".format(e))
    f.write("Exiting.\n")
    f.close()
    sys.exit(1)

# Create logging utility
LOGFILE = "/home/blake/Documents/logs/expense_report.log"
logging.basicConfig(filename=LOGFILE, level=logging.INFO)
logger = logging.getLogger("CashFlowParser")

# Pip requirements
try:
    # Import xlsx module
    from openpyxl import load_workbook

    # Import workbook path (private)
    from Workbook import *

    # Other utilities
    from tabulate import tabulate
    from dateutil.relativedelta import relativedelta
    from EmailUtils import FROM, TO, SERVER, PASSWORD, EMAILTXT
except Exception as e:
    logger.error("Couldn't import pip requirements:\n {}".format(e))
    logger.error("Make sure the requirements.txt file was properly imported.")
    logger.error("Exiting.")
    sys.exit(1)

# Class of functions that assist in sending
# the expense report.
class CashFlowParser:
    def __init__(self):
        try:
            logger.info("Initializing object...")
            # Workbook related items
            self.book = load_workbook(WBPATH)
            self.sheetnames = self.book.sheetnames

            # Relevant columns
            self.DATE = 1
            self.DESC = 2
            self.VALUE = 3

            logger.info("Initialization complete.")

        except Exception as e:
            logger.error("Failed to initialize.")
            logger.error("{}".format(e))
            logger.error("Exiting.")
            sys.exit(1)

    # Given: Excel formula (i.e.: =-1.25-8.24)
    # OR just a normal number - function will execute as normal
    # Returns the numerical value of the formula
    def numeric(self, formula):
        return eval(formula.replace("=", ""))

    # Return all applicable worksheets (date may span
    # multiple years)
    # Date must be datetime object
    def getSheets(self, date=None):
        if date is None:
            return None

        # List of sheets
        sheets = []
        # List of sheet names (for logging purposes)
        names = []
        # If the date spans multiple years (given date is not this year),
        # grab this years sheet first.
        if date.year != datetime.today().year:
            for name in self.sheetnames:
                if name == str(datetime.today().year):
                    sheets.append(self.book[name])
                    names.append(name)

        # Grab the given years sheet
        for name in self.sheetnames:
            if name == str(date.year):
                sheets.append(self.book[name])
                names.append(name)

        # Sheet list as string
        shs = ", ".join(names)
        logger.info("Reading " + str(len(names)) + " sheet(s): " + shs)

        return sheets

    # Takes given date, and what to order on
    # Finds all expenses within the given timeframe, sums them
    # and returns a sorted list
    def orderExpenses(self, date=None, order=None):
        logger.info("Ordering expenses...")
        if date is None or order is None:
            return None

        # To store data
        expenses = defaultdict(float)
        frequency = defaultdict(int)

        # Relevant worksheets
        # Could be > 1 (see getSheets)
        sheets = self.getSheets(date)

        # Collect data for all 1+ sheets
        print("Collecting data for " + str(len(sheets)) + " sheets")
        for sheet in sheets:
            logger.info("Collecting information from " + str(sheet))

            date_column = list(sheet.columns)[self.DATE]
            desc_column = list(sheet.columns)[self.DESC]
            value_column = list(sheet.columns)[self.VALUE]

            # Find row corresponding to given date
            start_row = None
            end_row = None
            print("Looking for a date after " + str(date))
            for cell in date_column[2:]:                    # Skip first 2 cells
                #print("Comparing " + str(cell.value))
                # "is not str": Ignore monthly summary titles
                # "start_row is None": Do not override when searching for end row
                # "cell.value is not None": Edge case
                if type(cell.value) is not str and start_row is None and cell.value is not None:
                    if cell.value >= date:
                        start_row = int(cell.coordinate[1:])    # Only need the row number
                        print("Found start row: " + str(start_row))
                if type(cell.value) is not str and cell.value is None:
                    end_row = int(cell.coordinate[1:])      # Only need the row number
                    break                                   # Finish at the first blank cell

            print("Starting at " + str(start_row))
            print("Ending at " + str(end_row))

            # Collect expenses
            if start_row is not None:
                for cell in desc_column[start_row:end_row-1]:
                    desc = cell.value
                    i = desc_column.index(cell)
                    val = value_column[i].value
                    if val is not None:
                        val = -self.numeric(str(val))
                        if val > 0:
                            expenses[desc] += val
                            frequency[desc] += 1

        ordered_list = sorted(expenses.items(), key=lambda kv: kv[1], reverse=True)
        final_list = []
        for exp in ordered_list:
            desc = exp[0]
            total = format(exp[1], '.2f')
            freq = frequency[desc]
            avg = '%.2f' % (exp[1]/freq)
            final_list.append((desc, freq, total, avg))

        logger.info("Sorted list was collected, returning result.")

        # Return first 10 items
        return sorted(final_list, key=lambda tup: float(tup[order]), reverse=True)[:10]

    # Create an HTML of all expenses, ordered by
    # highest to lowest gross spending, for the last
    # 6 months. Also returns a total sum of expenses for the last 6 months
    def sixMonthExpenses(self):
        logger.info("Collecting six month expenses...")
        six_months = datetime.today() - relativedelta(months=6)

        # "Total" index in tuple
        TOTAL = 2

        expense_list = self.orderExpenses(date=six_months, order=TOTAL)

        expense_sum = 0
        for expense in expense_list:
            # Values were converted to strings for table formatting,
            # these have to be reverted for summing
            if float(expense[TOTAL]) > 0:
                expense_sum += float(expense[TOTAL])

        logger.info("Returning HTML table and total 6 month expense.")
        return expense_sum, tabulate(expense_list, headers=["Desc.", "Freq", "Total", "Avg"], tablefmt='html', floatfmt=".2f")

    def twoWeekExpenses(self):
        two_weeks = datetime.today() - relativedelta(weeks=2)

        # "Total" index in tuple
        TOTAL = 2

        expense_list = self.orderExpenses(date=two_weeks, order=TOTAL)

        return tabulate(expense_list, headers=["Desc.", "Freq", "Total", "Avg"], tablefmt='html', floatfmt=".2f")

    # Create an HTML table of expenses,
    # ordered from highest to lowest gross spending,
    # for the last 3 months
    def grossExpenses(self):
        logger.info("Collecting expenses based on gross total...")
        # 3 months ago
        three_months = datetime.today() - relativedelta(months=3)

        # Tuple position for ordering
        TOTAL = 2

        gross_list = self.orderExpenses(three_months, order=TOTAL)

        logger.info("Returning HTML table for gross expenses.")
        return tabulate(gross_list, headers=["Desc.", "Freq", "Total", "Avg"], tablefmt='html', floatfmt=".2f")

    # Create an HTML table of expenses,
    # ordered by highest to lowest frequency
    # for the last 3 months
    def freqExpenses(self):
        logger.info("Collecting expenses based on frequency...")
        # 3 months ago
        three_months = datetime.today() - relativedelta(months=3)

        # Tuple positions, for ordering
        FREQUENCY = 1

        frequency_list = self.orderExpenses(date=three_months, order=FREQUENCY)

        logger.info("Returning HTML table for frequent expenses.")
        return tabulate(frequency_list, headers=["Desc.", "Freq", "Total", "Avg"], tablefmt='html', floatfmt=".2f")

    # Send an email with the given contents
    # Content should be formatted as HTML
    def sendMail(self, content):
        logger.info("Sending email...")
        msg = MIMEMultipart("alternative", None, [MIMEText(content, 'html')])

        msg['Subject'] = "Weekly Expense Report"
        msg['From'] = FROM
        msg['To'] = TO
        server = smtplib.SMTP(SERVER)
        server.ehlo()
        server.starttls()
        server.login(TO, PASSWORD)    # Authenticate with the actual gmail account
        server.sendmail(FROM, TO, msg.as_string())
        server.quit()

        logger.info("Email sent successfully.")

    # Create the expense report and send it.
    def sendExpenseReport(self):
        logger.info("Beginning expense report creation.")
        frequency_table = self.freqExpenses()
        gross_expense_table = self.grossExpenses()
        six_month_sum, six_month_table = self.sixMonthExpenses()
        two_weeks_table = self.twoWeekExpenses()
        content = EMAILTXT.format(two_weeks_table, frequency_table, gross_expense_table, six_month_sum, six_month_table)

        self.sendMail(content=content)
        logger.info("Expense report sent.")
