# Import 'load_workbook' module from 'openpyxl'
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('./test.xlsx')

# Get the sheet names
print(wb.sheetnames)

# Get a shhet by name
sheet = wb['Sheet3']

# Print the sheet title
sheet.title

# Get currently active sheet
anotherSheet = wb.active

# Check 'anotherSheet'
anotherSheet

# retrieve the value of a certain cell
sheet['A1'].value

# Select element 'B2' of your sheet
c = sheet['B2']

# Retrieve the row number of your element
c.row

# Retrieve the column letter of your element
c.column

# Retrieve the coordiantes of the cell
c.coordinate

# Retrieve cell calue
sheet.cell(row=1, column=2).value

# Print out values in column 2
for i in range(1, 4):
    print(i, sheet.cell(row=i, column=2).value)

# Import relevant modules from 'openpyxl.utils'
from openpyxl.utils import get_column_letter, column_index_from_string

# Return 'A'
get_column_letter(1)

# Return '1'
column_index_from_string('A')

# Print row per row
for cellObj in sheet['A1':'C3']:
    for cell in cellObj:
        print(cell.coordinate, cell.value)
    print('--- END ---')

# Retrieve the maximum amount of rows
print(sheet.max_row)

# Retrueve the maximm amount of columns
sheet.max_column
